from rest_framework import status, generics, permissions
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.views import APIView
from django.db.models import Q, Count
from django.utils import timezone
from django.contrib.auth import get_user_model
from django.http import HttpResponse
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, date, time
import io
import os
from .models import Event, Song, EventParticipant, EventStats
from .serializers import (
    EventSerializer, EventCreateSerializer, EventUpdateSerializer,
    EventStatsSerializer, DashboardSerializer
)

User = get_user_model()



class EventListView(generics.ListCreateAPIView):
    """List and create events"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.request.method == 'POST':
            return EventCreateSerializer
        return EventSerializer
    
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        event = serializer.save()
        # Return the created event with all related data
        return_serializer = EventSerializer(event)
        return Response(return_serializer.data, status=status.HTTP_201_CREATED)
    
    def get_queryset(self):
        queryset = Event.objects.select_related('created_by').prefetch_related('songs', 'dress_details', 'participants__user')
        
        # Filter by status if provided
        status_filter = self.request.query_params.get('status')
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        
        # Filter by date range if provided
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        
        # Sorting functionality
        sort_by = self.request.query_params.get('sort_by', 'date_time')
        sort_order = self.request.query_params.get('sort_order', 'asc')
        
        if sort_by == 'date_time':
            # Sort by date first, then by time
            if sort_order == 'desc':
                queryset = queryset.order_by('-date', '-time')
            else:  # asc
                queryset = queryset.order_by('date', 'time')
        elif sort_by == 'date':
            if sort_order == 'desc':
                queryset = queryset.order_by('-date')
            else:
                queryset = queryset.order_by('date')
        elif sort_by == 'time':
            if sort_order == 'desc':
                queryset = queryset.order_by('-time')
            else:
                queryset = queryset.order_by('time')
        elif sort_by == 'created':
            if sort_order == 'desc':
                queryset = queryset.order_by('-created_at')
            else:
                queryset = queryset.order_by('created_at')
        else:
            # Default sorting by date and time (nearest first)
            queryset = queryset.order_by('date', 'time')
        
        return queryset
    
    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class EventDetailView(generics.RetrieveUpdateDestroyAPIView):
    """Event detail view"""
    serializer_class = EventSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        return Event.objects.select_related('created_by').prefetch_related('songs', 'participants__user')
    
    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return EventUpdateSerializer
        return EventSerializer
    
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop('partial', False)
        instance = self.get_object()
        serializer = self.get_serializer(instance, data=request.data, partial=partial)
        serializer.is_valid(raise_exception=True)
        event = serializer.save()
        # Return the updated event with all related data
        return_serializer = EventSerializer(event)
        return Response(return_serializer.data)


class EventStatusUpdateView(APIView):
    """Update event status"""
    permission_classes = [permissions.IsAuthenticated]
    
    def patch(self, request, pk):
        try:
            event = Event.objects.get(pk=pk)
            new_status = request.data.get('status')
            
            if new_status not in ['pending', 'confirmed', 'completed', 'cancelled']:
                return Response(
                    {'error': 'Invalid status'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Only admins and coordinators can update status
            if not (request.user.is_admin or request.user.is_coordinator):
                return Response(
                    {'error': 'Permission denied'},
                    status=status.HTTP_403_FORBIDDEN
                )
            
            event.status = new_status
            event.save()
            
            return Response({
                'message': 'Event status updated successfully',
                'event': EventSerializer(event).data
            })
            
        except Event.DoesNotExist:
            return Response(
                {'error': 'Event not found'},
                status=status.HTTP_404_NOT_FOUND
            )


class DashboardView(APIView):
    """Dashboard data endpoint"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        # Get or create stats
        stats = EventStats.get_or_create_stats()
        stats.update_stats()
        
        # Get upcoming event (nearest event to current time, including pending)
        # First try to get future events
        upcoming_event = Event.objects.filter(
            date__gte=timezone.now().date()
        ).order_by('date', 'time').first()
        
        # If no future events, get the most recent event (nearest to today)
        if not upcoming_event:
            upcoming_event = Event.objects.order_by('date', 'time').first()
        
        # Get recent events
        recent_events = Event.objects.select_related('created_by').prefetch_related('songs')[:5]
        
        # Get total users
        total_users = User.objects.count()
        
        dashboard_data = {
            'stats': EventStatsSerializer(stats).data,
            'upcoming_event': EventSerializer(upcoming_event).data if upcoming_event else None,
            'recent_events': EventSerializer(recent_events, many=True).data,
            'total_users': total_users
        }
        
        return Response(dashboard_data)


class EventStatsView(APIView):
    """Event statistics endpoint"""
    permission_classes = [permissions.IsAuthenticated]
    
    def get(self, request):
        # Get or create stats
        stats = EventStats.get_or_create_stats()
        stats.update_stats()
        
        return Response(EventStatsSerializer(stats).data)


class EventByStatusView(generics.ListAPIView):
    """Get events by status"""
    serializer_class = EventSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        status = self.kwargs.get('status')
        return Event.objects.filter(status=status).select_related('created_by').prefetch_related('songs', 'participants__user').order_by('-created_at')


class EventSearchView(generics.ListAPIView):
    """Search events"""
    serializer_class = EventSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        queryset = Event.objects.select_related('created_by').prefetch_related('songs', 'dress_details', 'participants__user')
        
        # Search by place
        place = self.request.query_params.get('place')
        if place:
            queryset = queryset.filter(place__icontains=place)
        
        # Search by day
        day = self.request.query_params.get('day')
        if day:
            queryset = queryset.filter(day__icontains=day)
        
        # Search by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        
        return queryset.order_by('-created_at')


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def upcoming_events_view(request):
    """Get upcoming events"""
    events = Event.objects.filter(
        date__gte=timezone.now().date(),
        status__in=['pending', 'confirmed']
    ).select_related('created_by').prefetch_related('songs').order_by('date', 'time')
    
    return Response(EventSerializer(events, many=True).data)


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def past_events_view(request):
    """Get past events"""
    events = Event.objects.filter(
        date__lt=timezone.now().date()
    ).select_related('created_by').prefetch_related('songs').order_by('-date', '-time')
    
    return Response(EventSerializer(events, many=True).data)


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def join_event_view(request, pk):
    """Join an event"""
    try:
        event = Event.objects.get(pk=pk)
        
        # Check if user is already a participant
        if EventParticipant.objects.filter(event=event, user=request.user).exists():
            return Response(
                {'error': 'Already joined this event'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Create participation
        participant = EventParticipant.objects.create(
            event=event,
            user=request.user,
            is_confirmed=False
        )
        
        return Response({
            'message': 'Successfully joined the event',
            'participant': {
                'id': participant.id,
                'user': request.user.get_full_name(),
                'joined_at': participant.joined_at,
                'is_confirmed': participant.is_confirmed
            }
        })
        
    except Event.DoesNotExist:
        return Response(
            {'error': 'Event not found'},
            status=status.HTTP_404_NOT_FOUND
        )


@api_view(['DELETE'])
@permission_classes([permissions.IsAuthenticated])
def leave_event_view(request, pk):
    """Leave an event"""
    try:
        event = Event.objects.get(pk=pk)
        participant = EventParticipant.objects.get(event=event, user=request.user)
        participant.delete()
        
        return Response({'message': 'Successfully left the event'})
        
    except Event.DoesNotExist:
        return Response(
            {'error': 'Event not found'},
            status=status.HTTP_404_NOT_FOUND
        )
    except EventParticipant.DoesNotExist:
        return Response(
            {'error': 'Not participating in this event'},
            status=status.HTTP_400_BAD_REQUEST
        )


@api_view(['GET'])
@permission_classes([permissions.IsAuthenticated])
def download_sample_excel(request):
    """Download sample Excel file for event import"""
    try:
        # Create a new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Events Import Template"
        
        # Define headers based on Event model
        headers = [
            'Day', 'Date', 'Time', 'Duration (minutes)', 'Place', 'Number of Participants',
            'Status', 'Meeting Time', 'Meeting Date', 'Place of Meeting', 'Vehicle',
            'Camera Man', 'Participation Type', 'Event Reason'
        ]
        
        # Add headers with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add sample data
        sample_data = [
            ['Friday', '2024-01-15', '18:00', '120', 'Masjid Al-Noor', '25',
             'pending', '17:30', '2024-01-15', 'Main Hall', 'Bus #123',
             'Ahmed Ali', 'Recitation', 'Weekly Quran Recitation Session'],
            ['Saturday', '2024-01-20', '19:30', '90', 'Community Center', '15',
             'confirmed', '19:00', '2024-01-20', 'Conference Room', 'Van #456',
             'Omar Hassan', 'Listening', 'Special Event for New Muslims']
        ]
        
        for row, data in enumerate(sample_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="events_import_template.xlsx"'
        
        # Save workbook to response
        wb.save(response)
        return response
        
    except Exception as e:
        return Response(
            {'error': f'Failed to generate sample Excel: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@permission_classes([permissions.IsAuthenticated])
def import_events_excel(request):
    """Import events from Excel file"""
    try:
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        excel_file = request.FILES['file']
        
        # Validate file type
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return Response(
                {'error': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Read Excel file
        try:
            df = pd.read_excel(excel_file)
        except Exception as e:
            return Response(
                {'error': f'Failed to read Excel file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate required columns
        required_columns = [
            'Day', 'Date', 'Time', 'Duration (minutes)', 'Place', 'Number of Participants'
        ]
        missing_columns = [col for col in required_columns if col not in df.columns]
        
        if missing_columns:
            return Response(
                {'error': f'Missing required columns: {", ".join(missing_columns)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Process events
        imported_count = 0
        errors = []
        
        with transaction.atomic():
            for index, row in df.iterrows():
                try:
                    # Parse and validate data
                    day = str(row['Day']).strip()
                    event_date = pd.to_datetime(row['Date']).date()
                    event_time = pd.to_datetime(row['Time']).time()
                    duration = int(row['Duration (minutes)'])
                    place = str(row['Place']).strip()
                    participants = int(row['Number of Participants'])
                    
                    # Optional fields
                    status = str(row.get('Status', 'pending')).strip().lower()
                    if status not in ['pending', 'confirmed', 'completed', 'cancelled']:
                        status = 'pending'
                    
                    meeting_time = None
                    if pd.notna(row.get('Meeting Time')):
                        meeting_time = pd.to_datetime(row['Meeting Time']).time()
                    
                    meeting_date = None
                    if pd.notna(row.get('Meeting Date')):
                        meeting_date = pd.to_datetime(row['Meeting Date']).date()
                    
                    place_of_meeting = str(row.get('Place of Meeting', '')).strip() or None
                    vehicle = str(row.get('Vehicle', '')).strip() or None
                    camera_man = str(row.get('Camera Man', '')).strip() or None
                    participation_type = str(row.get('Participation Type', '')).strip() or None
                    event_reason = str(row.get('Event Reason', '')).strip() or None
                    
                    # Create event
                    event = Event.objects.create(
                        day=day,
                        date=event_date,
                        time=event_time,
                        duration=duration,
                        place=place,
                        number_of_participants=participants,
                        status=status,
                        meeting_time=meeting_time,
                        meeting_date=meeting_date,
                        place_of_meeting=place_of_meeting,
                        vehicle=vehicle,
                        camera_man=camera_man,
                        participation_type=participation_type,
                        event_reason=event_reason,
                        created_by=request.user
                    )
                    
                    imported_count += 1
                    
                except Exception as e:
                    errors.append(f'Row {index + 2}: {str(e)}')
                    continue
        
        # Update statistics
        try:
            stats = EventStats.get_or_create_stats()
            stats.update_stats()
        except:
            pass  # Don't fail import if stats update fails
        
        response_data = {
            'message': f'Successfully imported {imported_count} events',
            'imported_count': imported_count,
            'total_rows': len(df)
        }
        
        if errors:
            response_data['errors'] = errors[:10]  # Limit to first 10 errors
            response_data['error_count'] = len(errors)
        
        return Response(response_data, status=status.HTTP_200_OK)
        
    except Exception as e:
        return Response(
            {'error': f'Import failed: {str(e)}'},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )